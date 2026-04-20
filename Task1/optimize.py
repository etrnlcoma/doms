import os
import sys
import time
import multiprocessing as mp

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import mujoco
from openpyxl import Workbook, load_workbook

from pymoo.core.problem import Problem
from pymoo.core.callback import Callback
from pymoo.algorithms.moo.nsga2 import NSGA2
from pymoo.operators.sampling.lhs import LHS
from pymoo.operators.crossover.sbx import SBX
from pymoo.operators.mutation.pm import PM
from pymoo.optimize import minimize

from model_builder import build_model, check_reachable

# Настройки
XML_PATH  = "model.xml"
N_WORKERS = max(1, mp.cpu_count() - 1)
T_MAX     = 5.0

X_BOUNDS  = (80,  200)
L1_BOUNDS = (100, 400)
L2_BOUNDS = (100, 400)

JUMP_FREQ = 1.5
JUMP_AMP  = 1.0
JUMP_BIAS = 0.5

POP_SIZE = 40
N_GEN    = 60

RESULTS_DIR = "results"
PLOTS_DIR   = os.path.join(RESULTS_DIR, "plots")
LOG_XLSX    = os.path.join(RESULTS_DIR, "optimization_log.xlsx")

PENALTY_ENERGY = 1e6
PENALTY_V      = 0.0

# Нога не должна быть полностью прямой и не должна быть почти сложена
GEOM_MARGIN_FRAC = 0.1
# Максимально допустимое расхождение сайтов стоп
MAX_FOOT_SEPARATION = 0.1
# Минимальный подъём корпуса относительно высоты в момент отрыва
MIN_JUMP_HEIGHT = 0.2
# Сколько продолжать симуляцию после отрыва, чтобы зафиксировать макисмальный подъем
POST_TAKEOFF_TIME = 0.5


def geom_ok(x, l1, l2):
    lo, hi = abs(l1 - l2), l1 + l2
    m = GEOM_MARGIN_FRAC * x
    return (lo + m) <= x <= (hi - m)


# === ОДИН РОЛЛАУТ ===
def run_simulation(x_mm, l1_mm, l2_mm):
    x  = x_mm  / 1000.0
    l1 = l1_mm / 1000.0
    l2 = l2_mm / 1000.0

    if not geom_ok(x, l1, l2):
        return (PENALTY_ENERGY, PENALTY_V)

    try:
        check_reachable(x, l1, l2)
    except ValueError:
        return (PENALTY_ENERGY, PENALTY_V)

    try:
        model = build_model(XML_PATH, x_attach=x, L1=l1, L2=l2)
    except Exception:
        return (PENALTY_ENERGY, PENALTY_V)

    data = mujoco.MjData(model)
    dt = model.opt.timestep

    mujoco.mj_resetData(model, data)
    mujoco.mj_forward(model, data)

    main_box_id = model.body("main_box").id
    rf_site_id  = model.site("right_foot_site").id
    lf_site_id  = model.site("left_foot_site").id

    energy = 0.0
    v_takeoff = 0.0
    z_at_takeoff = 0.0
    max_box_z_post = 0.0
    max_foot_sep = 0.0
    was_in_contact = False
    takeoff_detected = False
    takeoff_time = -1.0

    while data.time < T_MAX:
        u = JUMP_BIAS + JUMP_AMP * np.sin(2 * np.pi * JUMP_FREQ * data.time)
        data.ctrl[0] =  u
        data.ctrl[1] = -u

        mujoco.mj_step(model, data)

        tau = np.array([data.sensor("right_hip_torque").data[0],
                        data.sensor("left_hip_torque").data[0]])
        omega = np.array([data.sensor("right_hip_omega").data[0],
                          data.sensor("left_hip_omega").data[0]])
        energy += max(0.0, float(np.sum(tau * omega))) * dt

        lt = float(data.sensor("left_foot_touch").data[0])
        rt = float(data.sensor("right_foot_touch").data[0])
        in_contact = (lt + rt) > 1e-6

        vz = float(data.sensor("main_box_vz").data[0])
        box_z = float(data.xpos[main_box_id, 2])

        sep = float(np.linalg.norm(data.site_xpos[rf_site_id]
                                   - data.site_xpos[lf_site_id]))
        if sep > max_foot_sep:
            max_foot_sep = sep

        if in_contact:
            was_in_contact = True
        elif was_in_contact and not takeoff_detected and vz > 0.0:
            # Контакт потерян и корпус движется вверх
            v_takeoff = vz
            z_at_takeoff = box_z
            max_box_z_post = box_z
            takeoff_detected = True
            takeoff_time = float(data.time)

        # После отрыва трекаем максимум высоты в баллистической фазе
        if takeoff_detected:
            if box_z > max_box_z_post:
                max_box_z_post = box_z
            if (data.time - takeoff_time) > POST_TAKEOFF_TIME:
                break

    if not takeoff_detected:
        return (PENALTY_ENERGY, PENALTY_V)

    # Проверка, что equality - связь стоп держится
    if max_foot_sep > MAX_FOOT_SEPARATION:
        return (PENALTY_ENERGY, PENALTY_V)

    # Проверка, что после отрыва корпус реально поднялся, а не просто дёрнулся
    if (max_box_z_post - z_at_takeoff) < MIN_JUMP_HEIGHT:
        return (PENALTY_ENERGY, PENALTY_V)

    # Кинетическая энергия корпуса при отрыве не может превышать работу моторов.
    total_mass = float(np.sum(model.body_mass))
    kinetic = 0.5 * total_mass * v_takeoff ** 2
    if energy + 0.05 < kinetic:
        return (PENALTY_ENERGY, PENALTY_V)

    return (float(energy), float(v_takeoff))


# Оценка популяции
def evaluate_population(params_list, pool=None):
    if pool is None:
        return [run_simulation(*p) for p in params_list]
    return pool.starmap(run_simulation, params_list)


# Логи
HEADERS = ["gen", "X_mm", "L1_mm", "L2_mm", "energy", "v_takeoff", "is_pareto"]


def init_xlsx(filename):
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "log"
    ws.append(HEADERS)
    wb.save(filename)


def _pareto_mask(F):
    n = len(F)
    mask = np.ones(n, dtype=bool)
    for i in range(n):
        if not mask[i]:
            continue
        for j in range(n):
            if i == j:
                continue
            if np.all(F[j] <= F[i]) and np.any(F[j] < F[i]):
                mask[i] = False
                break
    return mask


def save_epoch_to_xlsx(generation, X, F, filename):
    pareto = _pareto_mask(F)
    wb = load_workbook(filename)
    ws = wb["log"]
    for i in range(len(X)):
        ws.append([
            int(generation),
            float(X[i, 0]), float(X[i, 1]), float(X[i, 2]),
            float(F[i, 0]), float(-F[i, 1]),
            bool(pareto[i]),
        ])
    wb.save(filename)


# Оптмизация
class Optimizer(Problem):
    def __init__(self, pool=None):
        super().__init__(
            n_var=3, n_obj=2,
            xl=np.array([X_BOUNDS[0], L1_BOUNDS[0], L2_BOUNDS[0]], dtype=float),
            xu=np.array([X_BOUNDS[1], L1_BOUNDS[1], L2_BOUNDS[1]], dtype=float),
        )
        self.pool = pool

    def _evaluate(self, X, out, *args, **kwargs):
        params = [(row[0], row[1], row[2]) for row in X]
        results = evaluate_population(params, pool=self.pool)
        energies = np.array([r[0] for r in results], dtype=float)
        v        = np.array([r[1] for r in results], dtype=float)
        # min energy, max v_takeoff  →  min(-v)
        out["F"] = np.column_stack([energies, -v])


class LogCallback(Callback):
    def __init__(self, filename):
        super().__init__()
        self.filename = filename
        self.gen = 0
        self.t0 = time.time()

    def notify(self, algorithm):
        self.gen += 1
        pop = algorithm.pop
        X = pop.get("X")
        F = pop.get("F")
        save_epoch_to_xlsx(self.gen, X, F, self.filename)
        best_v = float(-F[:, 1].min())
        best_e = float(F[:, 0].min())
        dt = time.time() - self.t0
        print(f"[gen {self.gen:3d}/{algorithm.n_gen}]  "
              f"best v_takeoff={best_v:+.3f}  best energy={best_e:.2f}  "
              f"elapsed={dt:.1f}s")


def run_optimization(pop_size=POP_SIZE, n_gen=N_GEN, seed=0, use_pool=True):
    os.makedirs(RESULTS_DIR, exist_ok=True)
    init_xlsx(LOG_XLSX)

    pool = mp.Pool(N_WORKERS) if use_pool else None
    try:
        problem = Optimizer(pool=pool)
        algorithm = NSGA2(
            pop_size=pop_size,
            sampling=LHS(),
            crossover=SBX(prob=0.9, eta=15),
            mutation=PM(eta=20),
            eliminate_duplicates=True,
        )
        callback = LogCallback(LOG_XLSX)
        res = minimize(
            problem, algorithm,
            ("n_gen", n_gen),
            seed=seed,
            callback=callback,
            verbose=False,
        )
    finally:
        if pool is not None:
            pool.close()
            pool.join()

    return res


# Графики
def plot_results(log_xlsx, plots_dir):
    os.makedirs(plots_dir, exist_ok=True)
    df = pd.read_excel(log_xlsx, sheet_name="log")
    # отбрасываем штрафные (не было отрыва)
    df_ok = df[df["energy"] < PENALTY_ENERGY / 2].copy()

    last_gen = df["gen"].max()
    last = df[df["gen"] == last_gen]
    pf = last[last["is_pareto"]]
    plt.figure(figsize=(7, 5))
    plt.scatter(df_ok["energy"], df_ok["v_takeoff"], s=10, alpha=0.3, label="all")
    plt.scatter(pf["energy"], pf["v_takeoff"], s=40, c="red", label="pareto (last gen)")
    plt.xlabel("energy, J")
    plt.ylabel("v_takeoff, m/s")
    plt.title("Pareto front")
    plt.legend()
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(os.path.join(plots_dir, "pareto_front.png"), dpi=120)
    plt.close()

    by_gen_v = df_ok.groupby("gen")["v_takeoff"].max()
    plt.figure(figsize=(7, 4))
    plt.plot(by_gen_v.index, by_gen_v.values, marker="o", ms=3)
    plt.xlabel("generation")
    plt.ylabel("best v_takeoff, m/s")
    plt.title("Convergence: v_takeoff")
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(os.path.join(plots_dir, "convergence_v.png"), dpi=120)
    plt.close()

    by_gen_e = df_ok.groupby("gen")["energy"].min()
    plt.figure(figsize=(7, 4))
    plt.plot(by_gen_e.index, by_gen_e.values, marker="o", ms=3, color="tab:orange")
    plt.xlabel("generation")
    plt.ylabel("best (min) energy, J")
    plt.title("Convergence: energy")
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(os.path.join(plots_dir, "convergence_energy.png"), dpi=120)
    plt.close()

    x_med = df_ok["X_mm"].median()
    band = df_ok[np.abs(df_ok["X_mm"] - x_med) <= 15.0]
    if len(band) >= 10:
        plt.figure(figsize=(7, 5))
        sc = plt.scatter(band["L1_mm"], band["L2_mm"], c=band["v_takeoff"],
                         s=40, cmap="viridis")
        plt.colorbar(sc, label="v_takeoff, m/s")
        plt.xlabel("L1, mm")
        plt.ylabel("L2, mm")
        plt.title(f"v_takeoff at X≈{x_med:.0f} mm")
        plt.grid(True, alpha=0.3)
        plt.tight_layout()
        plt.savefig(os.path.join(plots_dir, "heatmap_L1_L2.png"), dpi=120)
        plt.close()


# Выбор лучшего компромиса по заданному отношению
def pick_best_compromise(res, weights=(0.0, 1.0)):
    # res.F — pareto front (последняя, отфильтрованная)
    # res.X — соответствующие параметры
    F = res.F.copy()
    # нормируем каждую колонку в [0,1]
    fmin = F.min(axis=0)
    fmax = F.max(axis=0)
    span = np.where(fmax - fmin > 1e-9, fmax - fmin, 1.0)
    Fn = (F - fmin) / span
    ideal = np.zeros(2)
    d = np.sqrt(((Fn - ideal) ** 2 * np.array(weights)).sum(axis=1))
    idx = int(np.argmin(d))
    x_mm, l1_mm, l2_mm = res.X[idx]
    energy = float(res.F[idx, 0])
    v      = float(-res.F[idx, 1])
    return (x_mm, l1_mm, l2_mm, energy, v)


# Визуализация лучшего
def visualize_best(x_mm, l1_mm, l2_mm):
    import mujoco.viewer
    model = build_model(
        XML_PATH,
        x_attach=x_mm / 1000.0, L1=l1_mm / 1000.0, L2=l2_mm / 1000.0,
    )
    data = mujoco.MjData(model)
    mujoco.mj_resetData(model, data)

    with mujoco.viewer.launch_passive(model, data) as viewer:
        viewer.opt.flags[mujoco.mjtVisFlag.mjVIS_JOINT] = True
        while viewer.is_running():
            step_start = time.time()
            u = JUMP_BIAS + JUMP_AMP * np.sin(2 * np.pi * JUMP_FREQ * data.time)
            data.ctrl[0] =  u
            data.ctrl[1] = -u
            mujoco.mj_step(model, data)
            viewer.sync()
            elapsed = time.time() - step_start
            time.sleep(max(0.0, model.opt.timestep - elapsed))

def main():
    mode = sys.argv[1] if len(sys.argv) > 1 else "full"

    res = run_optimization()
    plot_results(LOG_XLSX, PLOTS_DIR)

    x_mm, l1_mm, l2_mm, energy, v = pick_best_compromise(res)
    print("\n Лучший компромисс")
    print(f"X  = {x_mm:.2f} mm")
    print(f"L1 = {l1_mm:.2f} mm")
    print(f"L2 = {l2_mm:.2f} mm")
    print(f"energy = {energy:.3f} J")
    print(f"v_takeoff = {v:.3f} m/s")
    print(f"Лог: {LOG_XLSX}")
    print(f"Графики: {PLOTS_DIR}/")

    visualize_best(x_mm, l1_mm, l2_mm)


if __name__ == "__main__":
    main()
